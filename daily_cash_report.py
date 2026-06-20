"""
일일 시재자금보고 자동화 스크립트
- IBK기업은행 / 하나은행 CMS 거래내역 엑셀을 읽어
  시재자금현황_2026.xlsx / 시재자금현황sw_2026.xlsx 에 날짜 시트를 자동 생성·갱신합니다.
"""

from __future__ import annotations

import copy
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# 설정
# ──────────────────────────────────────────────
IBK_HEADER_MARKER = "사업장"          # IBK 파일 헤더에 반드시 존재하는 컬럼명
HANA_HEADER_MARKER = "기업명"         # 하나은행 파일 헤더에 반드시 존재하는 컬럼명

OUTPUT_IBK  = "시재자금현황_2026.xlsx"
OUTPUT_HANA = "시재자금현황sw_2026.xlsx"

# ── 보통예금 시재 현황 표 (A~J열) ──────────────────
COL_GUBUN        = 1   # A: 구분
COL_BIZNAME      = 2   # B: 사업장명
COL_ACCT_ALIAS   = 3   # C: 구좌명
COL_BANK         = 4   # D: 은행
COL_ACCT_NO      = 5   # E: 계좌번호
COL_PREV_BAL     = 6   # F: 전일잔액
COL_INCREASE     = 7   # G: 증가(입금합)
COL_DECREASE     = 8   # H: 감소(출금합)
COL_CURR_BAL     = 9   # I: 당일잔액
COL_NOTE         = 10  # J: 비고

BAL_TABLE_START_ROW = 25   # 25행: 첫 데이터 행 (24행은 헤더)
BAL_TABLE_END_ROW   = 92   # 93행이 합계 행이므로 데이터는 92행까지

# ── 당일 거래내역 표 (K~R열, 우측 블록) ────────────
TXN_COL_BIZPLACE = 11  # K: 사업장
TXN_COL_ALIAS    = 12  # L: 계좌별칭
TXN_COL_ACCTNO   = 13  # M: 계좌번호
TXN_COL_DATE     = 14  # N: 일자
TXN_COL_DEPOSIT  = 15  # O: 입금액
TXN_COL_WITHDRAW = 16  # P: 출금액
TXN_COL_DESC     = 17  # Q: 적요
TXN_COL_DETAIL   = 18  # R: 세부내용

TXN_DATA_START_ROW = 25  # 25행: 첫 데이터 행 (24행은 헤더)

# 소계/합계 행 감지에 사용할 B열 키워드
BAL_SKIP_KEYWORDS = ("소계", "합계", "계")


# ──────────────────────────────────────────────
# CMS 파일 파싱
# ──────────────────────────────────────────────

def detect_cms_type(path: Path) -> str:
    """헤더 컬럼을 보고 IBK / HANA 구분"""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        # .xls 는 openpyxl 미지원 → xlrd 로 재시도
        import xlrd
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
        if any(IBK_HEADER_MARKER in h for h in headers):
            return "IBK"
        if any(HANA_HEADER_MARKER in h for h in headers):
            return "HANA"
        raise ValueError(f"파일 형식을 판별할 수 없습니다: {path.name}")

    ws = wb.active
    headers = [str(cell.value or "") for cell in next(ws.iter_rows(max_row=1))]
    wb.close()
    if any(IBK_HEADER_MARKER in h for h in headers):
        return "IBK"
    if any(HANA_HEADER_MARKER in h for h in headers):
        return "HANA"
    raise ValueError(f"파일 형식을 판별할 수 없습니다: {path.name}")


def _read_rows(path: Path) -> tuple[list[str], list[list]]:
    """헤더 + 데이터 행 목록 반환. .xls/.xlsx 모두 지원"""
    suffix = path.suffix.lower()
    if suffix == ".xls":
        import xlrd
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
        rows = []
        for r in range(1, sheet.nrows):
            rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
        return headers, rows
    else:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return [], []
        headers = [str(v or "") for v in all_rows[0]]
        rows = [list(r) for r in all_rows[1:] if any(v is not None for v in r)]
        return headers, rows


def parse_ibk(path: Path) -> list[dict]:
    """IBK CMS 파일 → 거래 목록 (dict list)"""
    headers, rows = _read_rows(path)
    h = {name.strip(): i for i, name in enumerate(headers)}

    def col(row, name):
        idx = h.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    records = []
    for row in rows:
        date_val = col(row, "거래일자")
        if not date_val:
            continue
        records.append({
            "date":        _parse_date(date_val),
            "biz_place":   str(col(row, "사업장") or "").strip(),
            "acct_alias":  str(col(row, "계좌별칭") or "").strip(),
            "acct_no":     _clean_acct(col(row, "계좌번호")),
            "deposit":     _to_num(col(row, "입금액")),
            "withdraw":    _to_num(col(row, "출금액")),
            "desc":        str(col(row, "적요") or "").strip(),
            "bank":        str(col(row, "은행") or "").strip(),
            "balance":     _to_num(col(row, "잔액")),
            "currency":    str(col(row, "통화") or "KRW").strip(),
        })
    return records


def parse_hana(path: Path) -> list[dict]:
    """하나은행 CMS 파일 → 거래 목록 (dict list)"""
    headers, rows = _read_rows(path)
    h = {name.strip(): i for i, name in enumerate(headers)}

    def col(row, name):
        idx = h.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    records = []
    for row in rows:
        date_val = col(row, "거래일자")
        if not date_val:
            continue
        records.append({
            "date":        _parse_date(date_val),
            "biz_place":   str(col(row, "기업명") or "").strip(),
            "acct_alias":  str(col(row, "계좌별칭") or "").strip(),
            "acct_no":     _clean_acct(col(row, "계좌")),
            "deposit":     _to_num(col(row, "입금액")),
            "withdraw":    _to_num(col(row, "출금액")),
            "desc":        str(col(row, "적요") or "").strip(),
            "bank":        str(col(row, "은행") or "").strip(),
            "balance":     _to_num(col(row, "거래후잔액")),
            "currency":    str(col(row, "통화") or "KRW").strip(),
        })
    return records


# ──────────────────────────────────────────────
# 유틸리티
# ──────────────────────────────────────────────

def _parse_date(val) -> str:
    """거래일자를 'M.D' 형식 문자열로 변환 (예: '6.18')"""
    if val is None:
        return ""
    s = str(val).strip().replace("-", "").replace("/", "").replace(".", "")
    # 숫자 8자리 YYYYMMDD
    s = re.sub(r"\D", "", s)
    if len(s) == 8:
        m, d = int(s[4:6]), int(s[6:8])
        return f"{m}.{d}"
    return str(val).strip()


def _clean_acct(val) -> str:
    if val is None:
        return ""
    return re.sub(r"[\s\-]", "", str(val)).strip()


def _to_num(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = re.sub(r"[,\s]", "", str(val))
    try:
        return float(s)
    except ValueError:
        return 0.0


def _is_foreign_acct(ws, row: int) -> bool:
    """당일잔액 셀 값이 소수점이 있거나 외화 표시인지 간단 확인"""
    val = ws.cell(row=row, column=COL_CURR_BAL).value
    if val is None:
        return False
    s = str(val)
    # 소수점 있거나 숫자 아닌 문자(통화코드 등) 있으면 외화
    return bool(re.search(r"[A-Za-z]", s)) or ("." in s and not s.endswith(".0"))


def _sheet_date_key(name: str) -> tuple:
    """시트명 'M.D' → (M, D) 정수 튜플 (정렬용)"""
    m = re.match(r"(\d+)\.(\d+)", name)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    return (0, 0)


def _latest_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    """날짜 형식 시트 중 가장 최신 시트 반환"""
    date_sheets = [(name, _sheet_date_key(name)) for name in wb.sheetnames
                   if _sheet_date_key(name) != (0, 0)]
    if not date_sheets:
        raise RuntimeError("날짜 형식(M.D) 시트를 찾을 수 없습니다.")
    date_sheets.sort(key=lambda x: x[1])
    return wb[date_sheets[-1][0]]


# ──────────────────────────────────────────────
# 시트 복사 (서식 보존)
# ──────────────────────────────────────────────

def copy_sheet(wb: openpyxl.Workbook, src_ws, new_name: str):
    """src_ws 를 복사하여 new_name 시트를 wb 마지막에 추가"""
    tgt = wb.copy_worksheet(src_ws)
    tgt.title = new_name
    # copy_worksheet 는 탭이 소스 바로 뒤에 생기므로 마지막으로 이동
    wb.move_sheet(tgt, offset=len(wb.sheetnames) - 1 - wb.sheetnames.index(new_name))
    return tgt


# ──────────────────────────────────────────────
# 보통예금 시재 현황 표 처리
# ──────────────────────────────────────────────

def scan_balance_table(ws) -> dict[str, dict]:
    """
    보통예금 시재 현황 표를 스캔하여
    {clean_acct_no: {"row": int, "alias": str, "note": str, ...}} 반환.
    소계·합계 행은 건너뜀.
    """
    result = {}
    for row in range(BAL_TABLE_START_ROW, BAL_TABLE_END_ROW + 1):
        # 소계/합계 행(B열에 키워드) 건너뜀
        bizname_val = str(ws.cell(row=row, column=COL_BIZNAME).value or "").strip()
        if bizname_val in BAL_SKIP_KEYWORDS:
            continue

        acct_val = ws.cell(row=row, column=COL_ACCT_NO).value
        if not acct_val:
            continue
        acct_no = _clean_acct(acct_val)
        if not acct_no:
            continue
        alias = str(ws.cell(row=row, column=COL_ACCT_ALIAS).value or "").strip()
        note  = str(ws.cell(row=row, column=COL_NOTE).value or "").strip()
        result[acct_no] = {
            "row":        row,
            "alias":      alias,
            "note":       note,
            "is_project": "[과제]" in alias,
        }
    return result


def update_balance_table(ws, records: list[dict], balance_map: dict[str, dict]) -> list[str]:
    """
    거래내역으로 보통예금 시재 현황 표의 잔액을 갱신.
    반환: 경고 메시지 목록
    """
    warnings = []

    # 계좌별 집계: {acct_no: {"deposit": sum, "withdraw": sum, "last_balance": float, "currency": str}}
    agg: dict[str, dict] = defaultdict(lambda: {"deposit": 0.0, "withdraw": 0.0,
                                                  "last_balance": 0.0, "currency": "KRW"})
    for r in records:
        a = r["acct_no"]
        agg[a]["deposit"]      += r["deposit"]
        agg[a]["withdraw"]     += r["withdraw"]
        agg[a]["last_balance"]  = r["balance"]   # 마지막 행이 최신 잔액
        agg[a]["currency"]      = r["currency"]

    for acct_no, totals in agg.items():
        if acct_no not in balance_map:
            warnings.append(f"[신규계좌] {acct_no} — 시재 현황 표에 없는 계좌입니다. 수동으로 추가하세요.")
            continue

        info = balance_map[acct_no]
        row  = info["row"]

        # 외화 계좌 판별
        if _is_foreign_acct(ws, row) or totals["currency"] not in ("KRW", ""):
            warnings.append(
                f"[외화계좌] {acct_no} ({info['alias']}) — "
                f"통화={totals['currency']}, 입금={totals['deposit']}, 출금={totals['withdraw']}. "
                "외화 처리는 수동으로 확인하세요."
            )
            continue

        prev_bal = _to_num(ws.cell(row=row, column=COL_PREV_BAL).value)
        dep      = totals["deposit"]
        wdw      = totals["withdraw"]
        calc_bal = prev_bal + dep - wdw

        # [과제] 계좌 특별 처리
        if info["is_project"] and info["note"] != "포인트":
            # "일반과제" 행: 정상 갱신
            ws.cell(row=row, column=COL_INCREASE).value  = dep  if dep  else None
            ws.cell(row=row, column=COL_DECREASE).value  = wdw  if wdw  else None
            ws.cell(row=row, column=COL_CURR_BAL).value  = calc_bal

            # "포인트" 행 찾아서 역방향 반영
            pointer_row = _find_project_pointer_row(ws, acct_no, balance_map)
            if pointer_row and dep > 0:
                ptr_prev = _to_num(ws.cell(row=pointer_row, column=COL_PREV_BAL).value)
                ptr_curr = ptr_prev - dep
                ws.cell(row=pointer_row, column=COL_DECREASE).value = dep
                ws.cell(row=pointer_row, column=COL_CURR_BAL).value = ptr_curr
        else:
            ws.cell(row=row, column=COL_INCREASE).value  = dep  if dep  else None
            ws.cell(row=row, column=COL_DECREASE).value  = wdw  if wdw  else None
            ws.cell(row=row, column=COL_CURR_BAL).value  = calc_bal

        # 잔액 검증
        cms_bal = totals["last_balance"]
        if cms_bal and abs(calc_bal - cms_bal) > 1:
            warnings.append(
                f"[잔액불일치] {acct_no} ({info['alias']}): "
                f"계산={calc_bal:,.0f} / CMS={cms_bal:,.0f}. "
                "입력 누락 또는 중복 가능성을 확인하세요."
            )

    return warnings


def _find_project_pointer_row(ws, acct_no: str, balance_map: dict) -> Optional[int]:
    """같은 계좌번호에서 비고='포인트' 행의 row 번호 반환"""
    for info in balance_map.values():
        if (info.get("note") == "포인트"
                and info["is_project"]
                and _clean_acct(ws.cell(row=info["row"], column=COL_ACCT_NO).value) == acct_no):
            return info["row"]
    return None


# ──────────────────────────────────────────────
# 전일잔액 초기화 (신규 시트: 전일잔액 ← 이전 시트 당일잔액)
# ──────────────────────────────────────────────

def carry_forward_balances(prev_ws, new_ws, balance_map: dict):
    """이전 시트의 당일잔액을 새 시트의 전일잔액으로 복사"""
    for row in range(BAL_TABLE_START_ROW, BAL_TABLE_END_ROW + 1):
        prev_curr = prev_ws.cell(row=row, column=COL_CURR_BAL).value
        if prev_curr is None:
            continue
        new_ws.cell(row=row, column=COL_PREV_BAL).value = prev_curr
        # 증가/감소/당일잔액 초기화
        new_ws.cell(row=row, column=COL_INCREASE).value  = None
        new_ws.cell(row=row, column=COL_DECREASE).value  = None
        new_ws.cell(row=row, column=COL_CURR_BAL).value  = prev_curr  # 거래 없으면 동일


# ──────────────────────────────────────────────
# 당일 거래내역 표 입력
# ──────────────────────────────────────────────

def fill_transaction_table(ws, records: list[dict]) -> list[str]:
    """
    우측 당일 거래내역 표(K~R열, 절대 열 번호)에 CMS 원본 데이터 입력.
    숨긴 행은 필요시 펼쳐서 입력 후 다시 숨김.
    반환: 경고 메시지 목록
    """
    warnings = []
    if not records:
        return warnings

    data_rows = _get_txn_rows(ws)
    if len(records) > len(data_rows):
        warnings.append(
            f"[거래내역] 거래 {len(records)}건 > 표 행 {len(data_rows)}개. "
            "넘치는 거래는 입력되지 않았습니다. 수동으로 행을 추가하세요."
        )

    for i, rec in enumerate(records):
        if i >= len(data_rows):
            break
        row, was_hidden = data_rows[i]
        if was_hidden:
            ws.row_dimensions[row].hidden = False

        ws.cell(row=row, column=TXN_COL_BIZPLACE).value = rec["biz_place"]
        ws.cell(row=row, column=TXN_COL_ALIAS).value    = rec["acct_alias"]
        ws.cell(row=row, column=TXN_COL_ACCTNO).value   = rec["acct_no"]
        ws.cell(row=row, column=TXN_COL_DATE).value     = rec["date"]
        ws.cell(row=row, column=TXN_COL_DEPOSIT).value  = rec["deposit"]  if rec["deposit"]  else None
        ws.cell(row=row, column=TXN_COL_WITHDRAW).value = rec["withdraw"] if rec["withdraw"] else None
        ws.cell(row=row, column=TXN_COL_DESC).value     = rec["desc"]

    # 사용한 행 이후 숨겨진 행은 다시 숨김
    for j in range(len(records), len(data_rows)):
        row, was_hidden = data_rows[j]
        if was_hidden:
            ws.row_dimensions[row].hidden = True

    return warnings


def _get_txn_rows(ws) -> list[tuple[int, bool]]:
    """거래내역 표 데이터 행 목록 (row_no, 원래_숨김여부). B열 합계/소계 행 직전까지."""
    rows = []
    for row in range(TXN_DATA_START_ROW, ws.max_row + 1):
        b_val = str(ws.cell(row=row, column=COL_BIZNAME).value or "").strip()
        if b_val in BAL_SKIP_KEYWORDS:
            break
        rd = ws.row_dimensions.get(row)
        hidden = rd.hidden if rd else False
        rows.append((row, hidden))
    return rows


# ──────────────────────────────────────────────
# 메인 처리 흐름
# ──────────────────────────────────────────────

def determine_sheet_name(records: list[dict]) -> str:
    """
    여러 날짜가 섞인 경우 → 마지막 날짜(최대값)로 시트명 결정
    """
    dates = [r["date"] for r in records if r["date"]]
    if not dates:
        raise ValueError("거래일자를 읽을 수 없습니다.")
    # 'M.D' 를 (M, D) 튜플로 비교
    def to_tuple(s):
        m = re.match(r"(\d+)\.(\d+)", s)
        return (int(m.group(1)), int(m.group(2))) if m else (0, 0)
    latest = max(dates, key=to_tuple)
    return latest


def process_output(output_path: Path, records: list[dict], label: str):
    """output_path 에 신규 시트를 생성하고 거래내역을 반영"""
    if not output_path.exists():
        print(f"  [오류] 파일이 없습니다: {output_path}")
        return

    wb = load_workbook(output_path)
    sheet_name = determine_sheet_name(records)

    if sheet_name in wb.sheetnames:
        print(f"  [{label}] '{sheet_name}' 시트가 이미 존재합니다. 덮어씁니다.")
        del wb[sheet_name]

    prev_ws = _latest_sheet(wb)
    new_ws  = copy_sheet(wb, prev_ws, sheet_name)

    print(f"  [{label}] '{prev_ws.title}' → '{sheet_name}' 시트 복사 완료")

    # 전일잔액 이월
    balance_map = scan_balance_table(new_ws)
    carry_forward_balances(prev_ws, new_ws, balance_map)

    # 거래내역 표 초기화 (이전 날짜 데이터 삭제)
    _clear_transaction_table(new_ws)

    # 잔액 갱신
    warnings = update_balance_table(new_ws, records, balance_map)

    # 당일 거래내역 입력
    warnings += fill_transaction_table(new_ws, records)

    wb.save(output_path)
    print(f"  [{label}] 저장 완료: {output_path}")

    if warnings:
        print(f"\n  ── {label} 경고/확인 필요 항목 ──")
        for w in warnings:
            print(f"  ⚠  {w}")


def _clear_transaction_table(ws):
    """당일 거래내역 표 데이터 셀 초기화 (값만, 서식 유지)"""
    for row in range(TXN_DATA_START_ROW, ws.max_row + 1):
        b_val = str(ws.cell(row=row, column=COL_BIZNAME).value or "").strip()
        if b_val in BAL_SKIP_KEYWORDS:
            break
        for col in (TXN_COL_BIZPLACE, TXN_COL_ALIAS, TXN_COL_ACCTNO, TXN_COL_DATE,
                    TXN_COL_DEPOSIT, TXN_COL_WITHDRAW, TXN_COL_DESC, TXN_COL_DETAIL):
            ws.cell(row=row, column=col).value = None


# ──────────────────────────────────────────────
# 진입점
# ──────────────────────────────────────────────

def main():
    # 스크립트 위치의 test_data 폴더를 자동 스캔
    test_data_dir = Path(__file__).parent / "test_data"
    if not test_data_dir.exists():
        print(f"[오류] test_data 폴더가 없습니다: {test_data_dir}")
        sys.exit(1)

    cms_files = [
        f for f in test_data_dir.glob("*.xls*")
        if "시재자금" not in f.name
        and "입출내역" not in f.name
        and not f.name.startswith("~$")
    ]

    if not cms_files:
        print(f"CMS 파일을 찾을 수 없습니다: {test_data_dir}")
        sys.exit(1)

    ibk_records:  list[dict] = []
    hana_records: list[dict] = []

    for path in cms_files:
        print(f"읽는 중: {path.name}")
        try:
            cms_type = detect_cms_type(path)
        except Exception as e:
            print(f"  [오류] {e}")
            continue

        if cms_type == "IBK":
            ibk_records += parse_ibk(path)
            print(f"  → IBK 거래 {len(ibk_records)}건 파싱 완료")
        else:
            hana_records += parse_hana(path)
            print(f"  → 하나은행 거래 {len(hana_records)}건 파싱 완료")

    print()

    if ibk_records:
        process_output(Path(OUTPUT_IBK), ibk_records, "애니캐스팅(IBK)")
    else:
        print(f"IBK 거래 없음 — {OUTPUT_IBK} 갱신 건너뜀")

    print()

    if hana_records:
        process_output(Path(OUTPUT_HANA), hana_records, "SW법인(하나)")
    else:
        print(f"하나은행 거래 없음 — {OUTPUT_HANA} 갱신 건너뜀")

    print("\n완료.")


if __name__ == "__main__":
    main()